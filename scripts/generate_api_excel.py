from __future__ import annotations

"""
Generate an Excel API reference for the Order Service.

Outputs docs/api_documents.xlsx with styled sheets:
- Endpoints
- Request/Response Schemas
- Database Tables
- Status Enums
"""

from pathlib import Path
from datetime import datetime
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "docs" / "api_documents.xlsx"


def apply_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4F81BD")
    thin = Side(style="thin", color="9CAAC3")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def apply_body_style(cell):
    thin = Side(style="thin", color="D0D7E2")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_table(
    ws,
    columns: Sequence[str],
    rows: Iterable[Sequence[str | int | float | None]],
    widths: Sequence[int] | None = None,
):
    ws.append(columns)
    for cell in ws[1]:
        apply_header_style(cell)

    for row in rows:
        ws.append(list(row))

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            apply_body_style(cell)

    ws.freeze_panes = "A2"
    if widths:
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    endpoints = [
        (
            "POST",
            "/orders",
            "Tạo đơn mới",
            "Body: OrderCreate",
            "- 400: Order code already exists\n- 422: Validation",
            "201: OrderOut",
        ),
        (
            "GET",
            "/orders",
            "Danh sách đơn",
            "Query: skip (int, default 0), limit (int, default 50)",
            "- 422: Validation",
            "200: List[OrderOut]",
        ),
        (
            "GET",
            "/orders/{order_id}",
            "Lấy đơn theo order_id (UUID)",
            "Path: order_id",
            "- 404: Order not found\n- 422: Validation",
            "200: OrderOut",
        ),
        (
            "GET",
            "/orders/by-code/{order_code}",
            "Lấy đơn theo orderCode",
            "Path: order_code",
            "- 404: Order not found\n- 422: Validation",
            "200: OrderOut",
        ),
        (
            "GET",
            "/orders/status/{order_code}",
            "Lấy trạng thái đơn (nhẹ)",
            "Path: order_code",
            "- 404: Order not found\n- 422: Validation",
            "200: OrderStatusOut",
        ),
        (
            "POST",
            "/orders/status-update",
            "External push cập nhật trạng thái",
            "Body: ExternalStatusUpdate (cần orderId hoặc orderCode)",
            "- 400: Missing id/code\n- 404: Order not found\n- 422: Validation",
            "200: OrderOut (sau cập nhật)",
        ),
        (
            "PATCH",
            "/orders/{order_id}",
            "Cập nhật trạng thái nội bộ",
            "Path: order_id; Body: OrderUpdate",
            "- 404: Order not found\n- 422: Validation",
            "200: OrderOut",
        ),
        (
            "DELETE",
            "/orders/{order_id}",
            "Xóa đơn",
            "Path: order_id",
            "- 404: Order not found\n- 422: Validation",
            "204: No Content",
        ),
    ]

    ws = wb.create_sheet("Endpoints")
    add_table(
        ws,
        ["Method", "Path", "Mô tả", "Request", "Errors", "Response"],
        endpoints,
        widths=[10, 32, 30, 38, 30, 22],
    )

    schemas = [
        (
            "Customer",
            "customerId (str, <=50), name (str), phone (str), email (EmailStr?)",
        ),
        (
            "OrderItemCreate",
            "productId (str), productName (str), quantity (int>0), unitPrice (>=0), totalPrice (>=0)",
        ),
        (
            "Pricing",
            "subTotal (>=0), shippingFee (>=0, default 0), discount (>=0, default 0), totalAmount (>=0), currency (str, default VND)",
        ),
        (
            "Address",
            "receiverName (str), receiverPhone (str), fullAddress (str)",
        ),
        (
            "Shipper",
            "shipperId (str?), name (str?), phone (str?), vehicleType (motorbike|car|truck?)",
        ),
        (
            "Shipping",
            "shippingOrderCode (str?), status (NOT_CREATED|CREATED|PICKED|DELIVERING|DELIVERED|FAILED|CANCELLED, default NOT_CREATED), address (Address), shipper (Shipper?), estimatedDeliveryTime (datetime?), deliveredAt (datetime?), failedReason (str?)",
        ),
        (
            "OrderCreate",
            "orderCode (str), customer (Customer), items (List[OrderItemCreate]), pricing (Pricing), shipping (Shipping), orderStatus (DRAFT|CONFIRMED|CANCELLED|COMPLETED, default CONFIRMED), paymentMethod (str?), paymentStatus (str, default PENDING)",
        ),
        (
            "OrderUpdate",
            "orderStatus?, paymentStatus?, shippingStatus? (NOT_CREATED|CREATED|PICKED|DELIVERING|DELIVERED|FAILED|CANCELLED), shippingOrderCode?, shipper?, estimatedDeliveryTime?, deliveredAt?, failedReason?",
        ),
        (
            "OrderOut",
            "id (UUID), orderCode, customer, items(List[OrderItemOut]), pricing, shipping, orderStatus, paymentMethod?, paymentStatus, createdAt, updatedAt",
        ),
        (
            "OrderStatusOut",
            "orderCode, orderStatus, paymentStatus, shippingStatus, shippingOrderCode?, customer, pricing, createdAt, updatedAt",
        ),
        (
            "ExternalStatusUpdate",
            "orderId (UUID?) hoặc orderCode (str?) (cần ít nhất một), orderStatus?, paymentStatus?, shippingStatus?, shippingOrderCode?, shipper?, estimatedDeliveryTime?, deliveredAt?, failedReason?",
        ),
    ]

    ws = wb.create_sheet("Request-Response Schemas")
    add_table(
        ws,
        ["Schema", "Fields"],
        schemas,
        widths=[26, 120],
    )

    db_tables = [
        (
            "orders",
            (
                "id (UUID, PK), order_code (unique, str<=50), customer_id (str<=50), "
                "customer_name (str<=255), customer_phone (str<=30), customer_email (str<=255, nullable), "
                "subtotal (float), shipping_fee (float, default 0), discount (float, default 0), "
                "total_amount (float), currency (str<=10, default VND), payment_method (str<=50, nullable), "
                "payment_status (str<=30, default PENDING), shipping_order_code (str<=100, nullable), "
                "shipping_status (str<=30, default NOT_CREATED), receiver_name (str<=255), "
                "receiver_phone (str<=30), receiver_address (str<=500), shipper (JSONB, nullable), "
                "estimated_delivery_time (datetime tz, nullable), delivered_at (datetime tz, nullable), "
                "failed_reason (str<=500, nullable), order_status (str<=30, default CONFIRMED), "
                "created_at (datetime tz, default now), updated_at (datetime tz, auto update)"
            ),
        ),
        (
            "order_items",
            (
                "id (int, PK), order_id (UUID FK -> orders.id, cascade delete), "
                "product_id (str<=50), product_name (str<=255), quantity (int), "
                "unit_price (float), total_price (float)"
            ),
        ),
    ]

    ws = wb.create_sheet("Database Tables")
    add_table(ws, ["Table", "Columns"], db_tables, widths=[18, 120])

    status_enums = [
        ("orderStatus", "DRAFT, CONFIRMED (default), CANCELLED, COMPLETED"),
        ("paymentStatus", "PENDING (default) + tùy chỉnh khác"),
        (
            "shippingStatus",
            "NOT_CREATED (default), CREATED, PICKED, DELIVERING, DELIVERED, FAILED, CANCELLED",
        ),
    ]

    ws = wb.create_sheet("Status Enums")
    add_table(ws, ["Field", "Giá trị hợp lệ"], status_enums, widths=[22, 80])

    ws = wb.create_sheet("Metadata")
    add_table(
        ws,
        ["Key", "Value"],
        [
            ("Generated At", datetime.utcnow().isoformat() + "Z"),
            ("Service", "order-service"),
            ("Version", "0.1.0"),
            ("Base URL", "http://<host>:8000"),
        ],
        widths=[20, 60],
    )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


def main():
    output = build_workbook()
    print(f"Excel API docs generated at: {output}")


if __name__ == "__main__":
    main()

