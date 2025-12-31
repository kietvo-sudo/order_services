import pathlib


def ensure_openpyxl():
    try:
        import openpyxl  # type: ignore  # noqa: F401
    except ImportError:
        import subprocess
        import sys

        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl", "-q"]
        )


def style_header(ws):
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit_columns(ws):
    from openpyxl.utils import get_column_letter

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
            except Exception:
                v = ""
            if len(v) > max_length:
                max_length = len(v)
        ws.column_dimensions[column].width = min(max_length + 2, 60)


def build_workbook(path: pathlib.Path):
    from openpyxl import Workbook

    wb = Workbook()

    # Sheet: library
    ws_lib = wb.active
    ws_lib.title = "library"
    ws_lib.append(["Package", "Version / Source", "Description"])
    libs = [
        ("python", "3.11 (conda)", "Runtime"),
        ("fastapi", "", "Web framework"),
        ("uvicorn", "", "ASGI server"),
        ("sqlalchemy", "", "ORM for PostgreSQL"),
        ("psycopg2-binary", "", "PostgreSQL driver"),
        ("pydantic", "", "Data validation / serialization"),
        ("pydantic-settings", "", "Env-based configuration"),
        ("python-dotenv", "", "Load .env files"),
        ("typing_extensions", "", "Typing helpers"),
    ]
    for name, version, note in libs:
        ws_lib.append([name, version, note])
    style_header(ws_lib)
    autofit_columns(ws_lib)

    # Sheet: database structure
    ws_db = wb.create_sheet("database_structure")
    ws_db.append(["Table", "Column", "Type", "Constraints / Notes"])
    orders_columns = [
        ("orders", "id", "uuid", "PK, default uuid4"),
        ("orders", "order_code", "varchar(50)", "Unique order code"),
        ("orders", "customer_id", "varchar(50)", "Business customer id"),
        ("orders", "customer_name", "varchar(255)", "Customer full name"),
        ("orders", "customer_phone", "varchar(30)", "Customer phone"),
        ("orders", "customer_email", "varchar(255)", "Nullable email"),
        ("orders", "subtotal", "numeric", "Order subtotal (items total)"),
        ("orders", "shipping_fee", "numeric", "Shipping fee, default 0"),
        ("orders", "discount", "numeric", "Total discount, default 0"),
        ("orders", "total_amount", "numeric", "Final amount to pay"),
        ("orders", "currency", "varchar(10)", "Default VND"),
        ("orders", "payment_method", "varchar(50)", "Nullable, e.g. COD, BANK_TRANSFER"),
        ("orders", "payment_status", "varchar(30)", "PENDING / PAID / FAILED ..."),
        ("orders", "shipping_order_code", "varchar(100)", "External shipping code"),
        ("orders", "shipping_status", "varchar(30)", "NOT_CREATED / CREATED / ..."),
        ("orders", "receiver_name", "varchar(255)", "Shipping receiver name"),
        ("orders", "receiver_phone", "varchar(30)", "Shipping receiver phone"),
        ("orders", "receiver_address", "varchar(500)", "Full shipping address"),
        ("orders", "shipper", "jsonb", "shipperId, name, phone, vehicleType"),
        ("orders", "estimated_delivery_time", "timestamptz", "Estimated delivery time"),
        ("orders", "delivered_at", "timestamptz", "Actual delivered time"),
        ("orders", "failed_reason", "varchar(500)", "Reason if FAILED"),
        ("orders", "order_status", "varchar(30)", "DRAFT / CONFIRMED / COMPLETED / CANCELLED"),
        ("orders", "created_at", "timestamptz", "Auto now()"),
        ("orders", "updated_at", "timestamptz", "Auto now() on update"),
    ]
    items_columns = [
        ("order_items", "id", "serial", "PK"),
        ("order_items", "order_id", "uuid", "FK to orders.id, cascade delete"),
        ("order_items", "product_id", "varchar(50)", "Product code"),
        ("order_items", "product_name", "varchar(255)", "Product name"),
        ("order_items", "quantity", "int", "> 0"),
        ("order_items", "unit_price", "numeric", "Price per unit"),
        ("order_items", "total_price", "numeric", "quantity * unit_price"),
    ]
    for row in orders_columns + items_columns:
        ws_db.append(row)
    style_header(ws_db)
    autofit_columns(ws_db)

    # Sheet: mock data (based on JSON sample)
    ws_mock = wb.create_sheet("mockdata")
    ws_mock.append(["Field", "Example value", "Notes"])
    mock = [
        ("_id", "ObjectId()", "Mongo-style id (not stored in Postgres)"),
        ("orderCode", "ORD_20251217001", "Unique order code"),
        ("customer.customerId", "CUS001", "Business customer id"),
        ("customer.name", "Nguyen Van A", ""),
        ("customer.phone", "0909xxxxxx", ""),
        ("customer.email", "a.nguyen@gmail.com", ""),
        ("items[0].productId", "P001", ""),
        ("items[0].productName", "T-shirt male", ""),
        ("items[0].quantity", "2", ""),
        ("items[0].unitPrice", "150000", ""),
        ("items[0].totalPrice", "300000", ""),
        ("pricing.subTotal", "300000", ""),
        ("pricing.shippingFee", "30000", ""),
        ("pricing.discount", "0", ""),
        ("pricing.totalAmount", "330000", ""),
        ("shipping.status", "NOT_CREATED", "NOT_CREATED / CREATED / PICKED / DELIVERING / DELIVERED / FAILED / CANCELLED"),
        ("shipping.address.receiverName", "Nguyen Van A", ""),
        ("shipping.address.receiverPhone", "0909xxxxxx", ""),
        ("shipping.address.fullAddress", "123 Nguyen Van Linh, District 7, Ho Chi Minh City", ""),
        ("shipping.shipper.vehicleType", "motorbike", "motorbike | car | truck"),
        ("orderStatus", "CONFIRMED", "DRAFT / CONFIRMED / COMPLETED / CANCELLED"),
        ("createdAt", "2025-12-17T03:30:00.000Z", "ISO 8601"),
        ("updatedAt", "2025-12-17T03:30:00.000Z", "ISO 8601"),
    ]
    for field, example, note in mock:
        ws_mock.append([field, example, note])
    style_header(ws_mock)
    autofit_columns(ws_mock)

    # Sheet: API list (detailed request/response)
    ws_api = wb.create_sheet("apilist")
    ws_api.append(
        [
            "Method",
            "Path",
            "Summary",
            "Request body schema",
            "Request body fields",
            "Query / Path params",
            "Response schema",
            "Response fields",
        ]
    )

    api_rows = [
        (
            "POST",
            "/orders",
            "Create a new order",
            "OrderCreate",
            "orderCode (string, required); "
            "customer (Customer: customerId, name, phone, email?); "
            "items (OrderItemCreate[]: productId, productName, quantity>0, unitPrice>=0, totalPrice>=0); "
            "pricing (Pricing: subTotal, shippingFee, discount, totalAmount, currency); "
            "shipping (Shipping: shippingOrderCode?, status, address, shipper?, estimatedDeliveryTime?, deliveredAt?, failedReason?); "
            "orderStatus (string, default CONFIRMED); "
            "paymentMethod? (string); "
            "paymentStatus (string, default PENDING)",
            "",
            "OrderOut (201)",
            "id (uuid); orderCode; customer (Customer); items (OrderItemOut[]); "
            "pricing (Pricing); shipping (Shipping); "
            "orderStatus; paymentMethod?; paymentStatus; createdAt; updatedAt",
        ),
        (
            "GET",
            "/orders",
            "List orders",
            "",
            "",
            "Query: skip (int, default 0), limit (int, default 50)",
            "List[OrderOut] (200)",
            "Array of OrderOut",
        ),
        (
            "GET",
            "/orders/{order_id}",
            "Get order by id",
            "",
            "",
            "Path: order_id (uuid)",
            "OrderOut (200)",
            "Same as OrderOut",
        ),
        (
            "GET",
            "/orders/by-code/{order_code}",
            "Get order by orderCode (for other systems)",
            "",
            "",
            "Path: order_code (string)",
            "OrderOut (200)",
            "Same as OrderOut",
        ),
        (
            "PATCH",
            "/orders/{order_id}",
            "Update order status/payment/shipping",
            "OrderUpdate",
            "orderStatus? (string); paymentStatus? (string); "
            "shippingStatus? (string); shippingOrderCode? (string); "
            "shipper? (Shipper: shipperId?, name?, phone?, vehicleType?); "
            "estimatedDeliveryTime? (datetime); deliveredAt? (datetime); failedReason? (string)",
            "Path: order_id (uuid)",
            "OrderOut (200)",
            "Updated OrderOut",
        ),
        (
            "DELETE",
            "/orders/{order_id}",
            "Delete an order",
            "",
            "",
            "Path: order_id (uuid)",
            "No content (204)",
            "",
        ),
        (
            "GET",
            "/healthz",
            "Health check",
            "",
            "",
            "",
            "JSON (200)",
            "status (string), service (string), env (string)",
        ),
    ]

    for row in api_rows:
        ws_api.append(row)
    style_header(ws_api)
    autofit_columns(ws_api)

    wb.save(path)


def main():
    ensure_openpyxl()
    root = pathlib.Path(__file__).resolve().parent.parent
    output_path = root / "order_service_overview.xlsx"
    build_workbook(output_path)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()


